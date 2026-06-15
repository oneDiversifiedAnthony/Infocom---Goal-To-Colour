"""Convert countries.json to an Excel spreadsheet."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(SCRIPT_DIR, "countries.json")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "countries.xlsx")


def rgb_to_hex(rgb):
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def luminance(rgb):
    """Return relative luminance to decide black/white font."""
    r, g, b = [c / 255.0 for c in rgb]
    return 0.299 * r + 0.587 * g + 0.114 * b


def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Countries"

    headers = [
        "Country", "Universe", "Channel",
        "Colour 1 Name", "Colour 2 Name", "Colour 3 Name",
        "Colour 1 (R)", "Colour 1 (G)", "Colour 1 (B)", "Colour 1",
        "Colour 2 (R)", "Colour 2 (G)", "Colour 2 (B)", "Colour 2",
        "Colour 3 (R)", "Colour 3 (G)", "Colour 3 (B)", "Colour 3",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    teams_sorted = sorted(data["teams"].items(), key=lambda x: x[1].get("trigger", {}).get("channel", 999))
    for row_idx, (name, info) in enumerate(teams_sorted, 2):
        colours = info.get("colours", [[], [], []])
        c1 = colours[0] if len(colours) > 0 else []
        c2 = colours[1] if len(colours) > 1 else []
        c3 = colours[2] if len(colours) > 2 else []

        trigger = info.get("trigger", {})

        row_data = [
            name,
            trigger.get("universe", ""),
            trigger.get("channel", ""),
            c1[3] if len(c1) > 3 else "", c2[3] if len(c2) > 3 else "", c3[3] if len(c3) > 3 else "",
            c1[0] if c1 else "", c1[1] if c1 else "", c1[2] if c1 else "", "",
            c2[0] if c2 else "", c2[1] if c2 else "", c2[2] if c2 else "", "",
            c3[0] if c3 else "", c3[1] if c3 else "", c3[2] if c3 else "", "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

        # Fill colour preview cells
        for colour, col_offset in [(c1, 10), (c2, 14), (c3, 18)]:
            if colour and len(colour) >= 3:
                cell = ws.cell(row=row_idx, column=col_offset)
                rgb = colour[:3]
                hex_color = rgb_to_hex(rgb)
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.value = f"#{hex_color}"
                font_color = "FFFFFF" if luminance(rgb) < 0.5 else "000000"
                cell.font = Font(color=font_color, size=9)
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
